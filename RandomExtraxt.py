
import openpyxl
import random
from openpyxl import load_workbook

#PATH #路径名
INPUT_FILES_BASE_PATH ="D:\\Users\\User\\Desktop\\"
EXCEL_FILENAME = 'name.xlsx'

def get_row_value(ws,row):
    col_num = ws.max_colum
    row_data = []
    for i in range(1,col_num+1):
        cell_value = ws.cell(row=row, column=i).value
        row_data.append(cell_value)
    return row_data

for i in range(1):
    
    #input #输入
    wb = load_workbook(r"{}".format(INPUT_FILES_BASE_PATH) + "{}".format(EXCEL_FILENAME))
    sheet = wb.active
    row_num = sheet.max_row
    #Take 10% of the sample at random, leaving the first row of table heads alone.
    # Variable function to achieve different random extraction rates
    # 随机抽取10%样本，第一行表头不取。可更改函数实现不同随机抽取率
    random_num = random.sample(range(2,row_num+1),row_num//10)

    #Write into new form #写入表格
    
    #RandomExtract part #随即提取的部分
    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active
    sheet2.append(get_row_value(sheet,1))
    for j in random_num:
        sheet2.append(get_row_value(sheet,j))
   
    #sheet2.append(['The random number generated is：'] + random_num)
    # If a random number needs to be generated, it is printed to the last row of the table
    #sheet2.append(['生成的随机数为：'] + random_num) 若需要生成的随机数，则会输出到表格最后一行

    #output #输出
    out_file_name1 = 'RandomExtract.xlsx'
    wb2.save(out_file_name1)

    # Remainder #剩下的部分
    wb3 = openpyxl.Workbook()
    sheet3 = wb3.active
    sheet3.append(get_row_value(sheet, 1))
    for m in range(2,row_num+1):
        if m not in random_num:
            sheet3.append(get_row_value(sheet, m))

    # output #输出
    out_file_name2 = 'RandomRemain.xlsx'
    wb2.save(out_file_name2)

    print('Success extract')

